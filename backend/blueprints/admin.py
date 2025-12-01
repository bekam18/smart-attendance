from flask import Blueprint, request, jsonify
from flask_jwt_extended import jwt_required
from bson import ObjectId
from datetime import datetime
import os

from db.mongo import get_db
from utils.security import hash_password, role_required
from config import config

admin_bp = Blueprint('admin', __name__)

@admin_bp.route('/add-instructor', methods=['POST'])
@jwt_required()
@role_required('admin')
def add_instructor():
    """Add a new instructor (admin only)"""
    data = request.get_json()
    
    required_fields = ['username', 'password', 'email', 'name', 'course_name', 'class_year']
    if not all(field in data for field in required_fields):
        return jsonify({'error': 'Missing required fields'}), 400
    
    # Validate session types - at least one must be selected
    lab_session = data.get('lab_session', False)
    theory_session = data.get('theory_session', False)
    
    if not lab_session and not theory_session:
        return jsonify({'error': 'At least one session type (Lab or Theory) must be selected'}), 400
    
    db = get_db()
    
    # Check if username or email already exists
    if db.users.find_one({'username': data['username']}):
        return jsonify({'error': 'Username already exists'}), 409
    
    if db.users.find_one({'email': data['email']}):
        return jsonify({'error': 'Email already exists'}), 409
    
    # Build session types array
    session_types = []
    if lab_session:
        session_types.append('lab')
    if theory_session:
        session_types.append('theory')
    
    # Create instructor user
    user_doc = {
        'username': data['username'],
        'password': hash_password(data['password']),
        'email': data['email'],
        'name': data['name'],
        'role': 'instructor',
        'department': data.get('department', ''),
        'course_name': data['course_name'],
        'class_year': str(data['class_year']),  # Ensure string format
        'session_types': session_types,  # ['lab'], ['theory'], or ['lab', 'theory']
        'created_at': datetime.utcnow()
    }
    
    result = db.users.insert_one(user_doc)
    
    print(f"✅ Instructor added: {data['name']} - Course: {data['course_name']} - Year: {data['class_year']} - Sessions: {session_types}")
    
    return jsonify({
        'message': 'Instructor added successfully',
        'instructor_id': str(result.inserted_id)
    }), 201

@admin_bp.route('/instructors', methods=['GET'])
@jwt_required()
@role_required('admin')
def get_instructors():
    """Get all instructors (admin only)"""
    db = get_db()
    
    instructors = db.users.find({'role': 'instructor'})
    
    instructor_list = []
    for instructor in instructors:
        instructor_list.append({
            'id': str(instructor['_id']),
            'username': instructor['username'],
            'name': instructor['name'],
            'email': instructor['email'],
            'department': instructor.get('department', ''),
            'course_name': instructor.get('course_name', ''),
            'class_year': instructor.get('class_year', ''),
            'session_types': instructor.get('session_types', []),
            'enabled': instructor.get('enabled', True),
            'created_at': instructor['created_at'].isoformat()
        })
    
    return jsonify(instructor_list), 200

@admin_bp.route('/add-student', methods=['POST'])
@jwt_required()
@role_required('admin')
def add_student():
    """Add a new student (admin only)"""
    data = request.get_json()
    
    print(f"🔍 Add student request: {data}")
    
    required_fields = ['username', 'password', 'email', 'name', 'student_id']
    if not all(field in data for field in required_fields):
        missing = [f for f in required_fields if f not in data]
        print(f"❌ Missing fields: {missing}")
        return jsonify({'error': f'Missing required fields: {", ".join(missing)}'}), 400
    
    db = get_db()
    
    # Check if username or email already exists
    if db.users.find_one({'username': data['username']}):
        print(f"❌ Username already exists: {data['username']}")
        return jsonify({'error': 'Username already exists'}), 409
    
    if db.users.find_one({'email': data['email']}):
        print(f"❌ Email already exists: {data['email']}")
        return jsonify({'error': 'Email already exists'}), 409
    
    if db.students.find_one({'student_id': data['student_id']}):
        print(f"❌ Student ID already exists: {data['student_id']}")
        return jsonify({'error': 'Student ID already exists'}), 409
    
    # Create user
    user_doc = {
        'username': data['username'],
        'password': hash_password(data['password']),
        'email': data['email'],
        'name': data['name'],
        'role': 'student',
        'created_at': datetime.utcnow()
    }
    
    user_result = db.users.insert_one(user_doc)
    user_id = str(user_result.inserted_id)
    
    # Create student profile
    student_doc = {
        'user_id': user_id,
        'student_id': data['student_id'],
        'name': data['name'],
        'email': data['email'],
        'department': data.get('department', ''),
        'year': data.get('year', ''),
        'face_registered': False,
        'created_at': datetime.utcnow()
    }
    
    db.students.insert_one(student_doc)
    
    print(f"✅ Student added successfully: {data['student_id']}")
    
    return jsonify({
        'message': 'Student added successfully',
        'student_id': data['student_id']
    }), 201

@admin_bp.route('/students', methods=['GET'])
@jwt_required()
@role_required('admin', 'instructor')
def get_students():
    """Get all students"""
    db = get_db()
    
    students = db.students.find()
    
    student_list = []
    for student in students:
        # Get user to check enabled status
        user = db.users.find_one({'_id': ObjectId(student['user_id'])}) if student.get('user_id') else None
        
        student_list.append({
            'id': str(student['_id']),
            'student_id': student['student_id'],
            'name': student['name'],
            'email': student['email'],
            'department': student.get('department', ''),
            'year': student.get('year', ''),
            'section': student.get('section', ''),
            'face_registered': student.get('face_registered', False),
            'enabled': user.get('enabled', True) if user else True,
            'created_at': student['created_at'].isoformat()
        })
    
    return jsonify(student_list), 200

@admin_bp.route('/attendance/all', methods=['GET'])
@jwt_required()
@role_required('admin')
def get_all_attendance():
    """Get all attendance records with advanced filters (admin only)"""
    db = get_db()
    
    # Get query parameters
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    student_id = request.args.get('student_id')
    section = request.args.get('section')
    instructor_id = request.args.get('instructor_id')
    
    # Build query
    query = {}
    
    # Date range filter
    if start_date or end_date:
        date_query = {}
        if start_date:
            date_query['$gte'] = start_date
        if end_date:
            date_query['$lte'] = end_date
        if date_query:
            query['date'] = date_query
    
    # Student filter
    if student_id:
        query['student_id'] = student_id
    
    # Section filter
    if section:
        query['section_id'] = section
    
    # Instructor filter
    if instructor_id:
        query['instructor_id'] = instructor_id
    
    print(f"📊 Admin fetching attendance with query: {query}")
    
    # Get attendance records
    attendance_records = list(db.attendance.find(query).sort('timestamp', -1).limit(1000))
    
    records = []
    for record in attendance_records:
        # Get student info
        student = db.students.find_one({'student_id': record['student_id']})
        
        # Get session info
        session = db.sessions.find_one({'_id': ObjectId(record.get('session_id', ''))}) if record.get('session_id') else None
        
        # Get instructor info
        instructor = db.users.find_one({'_id': ObjectId(record.get('instructor_id', ''))}) if record.get('instructor_id') else None
        
        records.append({
            'id': str(record['_id']),
            'student_id': record['student_id'],
            'student_name': student['name'] if student else 'Unknown',
            'section': student.get('section', record.get('section_id', 'N/A')) if student else record.get('section_id', 'N/A'),
            'instructor_name': instructor['name'] if instructor else 'Unknown',
            'instructor_id': record.get('instructor_id', ''),
            'session_id': record.get('session_id', ''),
            'session_name': session['name'] if session else 'Unknown',
            'timestamp': record['timestamp'].isoformat(),
            'date': record['date'],
            'confidence': record.get('confidence', 0),
            'status': record.get('status', 'present')
        })
    
    print(f"✅ Returning {len(records)} attendance records")
    return jsonify(records), 200


@admin_bp.route('/attendance/export/csv', methods=['GET'])
@jwt_required()
@role_required('admin')
def export_attendance_csv():
    """Export all attendance records to CSV (admin only)"""
    import csv
    import io
    from flask import send_file
    
    db = get_db()
    
    # Get same filters
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    student_id = request.args.get('student_id')
    section = request.args.get('section')
    instructor_id = request.args.get('instructor_id')
    
    # Build query (same as get_all_attendance)
    query = {}
    if start_date or end_date:
        date_query = {}
        if start_date:
            date_query['$gte'] = start_date
        if end_date:
            date_query['$lte'] = end_date
        if date_query:
            query['date'] = date_query
    if student_id:
        query['student_id'] = student_id
    if section:
        query['section_id'] = section
    if instructor_id:
        query['instructor_id'] = instructor_id
    
    # Get records
    records = list(db.attendance.find(query).sort('timestamp', -1))
    
    # Create CSV in memory
    output = io.StringIO()
    writer = csv.writer(output)
    
    # Write header
    writer.writerow(['Date', 'Time', 'Student ID', 'Student Name', 'Section', 'Instructor', 'Session', 'Confidence', 'Status'])
    
    # Write data
    for record in records:
        student = db.students.find_one({'student_id': record['student_id']})
        session = db.sessions.find_one({'_id': ObjectId(record.get('session_id', ''))}) if record.get('session_id') else None
        instructor = db.users.find_one({'_id': ObjectId(record.get('instructor_id', ''))}) if record.get('instructor_id') else None
        
        writer.writerow([
            record['date'],
            record['timestamp'].strftime('%H:%M:%S'),
            record['student_id'],
            student['name'] if student else 'Unknown',
            student.get('section', record.get('section_id', 'N/A')) if student else record.get('section_id', 'N/A'),
            instructor['name'] if instructor else 'Unknown',
            session['name'] if session else 'Unknown',
            f"{record.get('confidence', 0):.2%}",
            record.get('status', 'present')
        ])
    
    # Create response
    output.seek(0)
    csv_data = output.getvalue()
    csv_bytes = io.BytesIO(csv_data.encode('utf-8'))
    csv_bytes.seek(0)
    
    filename = f'all_attendance_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv'
    
    return send_file(
        csv_bytes,
        mimetype='text/csv',
        as_attachment=True,
        download_name=filename
    )


@admin_bp.route('/attendance/export/excel', methods=['GET'])
@jwt_required()
@role_required('admin')
def export_attendance_excel():
    """Export all attendance records to Excel (admin only)"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill
        import io
        from flask import send_file
        
        db = get_db()
        
        # Get same filters
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')
        student_id = request.args.get('student_id')
        section = request.args.get('section')
        instructor_id = request.args.get('instructor_id')
        
        # Build query
        query = {}
        if start_date or end_date:
            date_query = {}
            if start_date:
                date_query['$gte'] = start_date
            if end_date:
                date_query['$lte'] = end_date
            if date_query:
                query['date'] = date_query
        if student_id:
            query['student_id'] = student_id
        if section:
            query['section_id'] = section
        if instructor_id:
            query['instructor_id'] = instructor_id
        
        # Get records
        records = list(db.attendance.find(query).sort('timestamp', -1))
        
        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "All Attendance Records"
        
        # Style header
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        # Write header
        headers = ['Date', 'Time', 'Student ID', 'Student Name', 'Section', 'Instructor', 'Session', 'Confidence', 'Status']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
        
        # Write data
        for row_idx, record in enumerate(records, 2):
            student = db.students.find_one({'student_id': record['student_id']})
            session = db.sessions.find_one({'_id': ObjectId(record.get('session_id', ''))}) if record.get('session_id') else None
            instructor = db.users.find_one({'_id': ObjectId(record.get('instructor_id', ''))}) if record.get('instructor_id') else None
            
            ws.cell(row=row_idx, column=1, value=record['date'])
            ws.cell(row=row_idx, column=2, value=record['timestamp'].strftime('%H:%M:%S'))
            ws.cell(row=row_idx, column=3, value=record['student_id'])
            ws.cell(row=row_idx, column=4, value=student['name'] if student else 'Unknown')
            ws.cell(row=row_idx, column=5, value=student.get('section', record.get('section_id', 'N/A')) if student else record.get('section_id', 'N/A'))
            ws.cell(row=row_idx, column=6, value=instructor['name'] if instructor else 'Unknown')
            ws.cell(row=row_idx, column=7, value=session['name'] if session else 'Unknown')
            ws.cell(row=row_idx, column=8, value=f"{record.get('confidence', 0):.2%}")
            ws.cell(row=row_idx, column=9, value=record.get('status', 'present'))
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to BytesIO
        excel_bytes = io.BytesIO()
        wb.save(excel_bytes)
        excel_bytes.seek(0)
        
        filename = f'all_attendance_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        
        return send_file(
            excel_bytes,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
    
    except ImportError:
        return jsonify({'error': 'openpyxl not installed', 'message': 'Install with: pip install openpyxl'}), 500
    except Exception as e:
        print(f"❌ Error exporting Excel: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': 'Failed to export Excel', 'message': str(e)}), 500

@admin_bp.route('/upload-model', methods=['POST'])
@jwt_required()
@role_required('admin')
def upload_model():
    """Upload model files (admin only)"""
    if 'file' not in request.files:
        return jsonify({'error': 'No file provided'}), 400
    
    file = request.files['file']
    file_type = request.form.get('type', 'classifier')
    
    if file.filename == '':
        return jsonify({'error': 'No file selected'}), 400
    
    # Save file to models directory
    filename_map = {
        'classifier': 'face_classifier_v1.pkl',
        'encoder': 'label_encoder.pkl',
        'classes': 'label_encoder_classes.npy'
    }
    
    filename = filename_map.get(file_type, file.filename)
    filepath = os.path.join(config.MODEL_PATH, filename)
    
    file.save(filepath)
    
    return jsonify({
        'message': 'Model file uploaded successfully',
        'filename': filename
    }), 200

@admin_bp.route('/instructor/<instructor_id>', methods=['DELETE'])
@jwt_required()
@role_required('admin')
def delete_instructor(instructor_id):
    """Delete an instructor (admin only)"""
    db = get_db()
    
    try:
        result = db.users.delete_one({'_id': ObjectId(instructor_id), 'role': 'instructor'})
        
        if result.deleted_count == 0:
            return jsonify({'error': 'Instructor not found'}), 404
        
        print(f"✅ Instructor deleted: {instructor_id}")
        return jsonify({'message': 'Instructor deleted successfully'}), 200
        
    except Exception as e:
        print(f"❌ Error deleting instructor: {e}")
        return jsonify({'error': 'Failed to delete instructor'}), 500

@admin_bp.route('/student/<student_id>', methods=['DELETE'])
@jwt_required()
@role_required('admin')
def delete_student(student_id):
    """Delete a student (admin only)"""
    db = get_db()
    
    try:
        # Find student to get user_id
        student = db.students.find_one({'_id': ObjectId(student_id)})
        
        if not student:
            return jsonify({'error': 'Student not found'}), 404
        
        # Delete student record
        db.students.delete_one({'_id': ObjectId(student_id)})
        
        # Delete user record
        db.users.delete_one({'_id': ObjectId(student['user_id'])})
        
        # Delete attendance records
        db.attendance.delete_many({'student_id': student['student_id']})
        
        print(f"✅ Student deleted: {student_id}")
        return jsonify({'message': 'Student deleted successfully'}), 200
        
    except Exception as e:
        print(f"❌ Error deleting student: {e}")
        return jsonify({'error': 'Failed to delete student'}), 500

@admin_bp.route('/stats', methods=['GET'])
@jwt_required()
@role_required('admin')
def get_stats():
    """Get system statistics with optional date filter (admin only)"""
    from datetime import datetime as dt, timedelta
    
    db = get_db()
    
    # Get date parameter (default to today)
    date_param = request.args.get('date')
    
    if date_param:
        # Use provided date
        target_date = date_param
    else:
        # Default to today
        target_date = dt.now().strftime('%Y-%m-%d')
    
    # Calculate 12 hours ago from the target date
    target_datetime = dt.strptime(target_date, '%Y-%m-%d')
    twelve_hours_ago = target_datetime - timedelta(hours=12)
    
    print(f"📊 Getting stats for date: {target_date}")
    
    # Base stats (always total)
    stats = {
        'total_students': db.students.count_documents({}),
        'total_instructors': db.users.count_documents({'role': 'instructor'}),
        'students_with_face': db.students.count_documents({'face_registered': True}),
        'selected_date': target_date
    }
    
    # Date-filtered stats
    if date_param:
        # Specific date - show all records for that day
        stats['total_attendance_records'] = db.attendance.count_documents({'date': target_date})
        stats['active_sessions'] = db.sessions.count_documents({
            'status': 'active',
            'start_time': {
                '$gte': target_datetime,
                '$lt': target_datetime + timedelta(days=1)
            }
        })
    else:
        # Today - show last 12 hours
        stats['total_attendance_records'] = db.attendance.count_documents({
            'date': target_date,
            'timestamp': {'$gte': twelve_hours_ago}
        })
        stats['active_sessions'] = db.sessions.count_documents({'status': 'active'})
    
    print(f"✅ Stats: {stats}")
    return jsonify(stats), 200


@admin_bp.route('/settings', methods=['GET'])
@jwt_required()
@role_required('admin')
def get_admin_settings():
    """Get admin settings"""
    db = get_db()
    
    settings = db.admin_settings.find_one({})
    
    if not settings:
        # Return defaults
        return jsonify({
            'face_recognition_threshold': 0.60,
            'session_timeout_minutes': 120
        }), 200
    
    return jsonify({
        'face_recognition_threshold': settings.get('face_recognition_threshold', 0.60),
        'session_timeout_minutes': settings.get('session_timeout_minutes', 120)
    }), 200


@admin_bp.route('/settings', methods=['PUT'])
@jwt_required()
@role_required('admin')
def update_admin_settings():
    """Update admin settings"""
    try:
        data = request.get_json()
        db = get_db()
        
        settings_doc = {
            'face_recognition_threshold': float(data.get('face_recognition_threshold', 0.60)),
            'session_timeout_minutes': int(data.get('session_timeout_minutes', 120)),
            'updated_at': datetime.utcnow()
        }
        
        db.admin_settings.update_one(
            {},
            {'$set': settings_doc},
            upsert=True
        )
        
        print(f"✅ Admin settings updated: {settings_doc}")
        return jsonify({'message': 'Settings updated successfully'}), 200
    
    except Exception as e:
        print(f"❌ Error updating settings: {e}")
        return jsonify({'error': 'Failed to update settings', 'message': str(e)}), 500


@admin_bp.route('/active-sessions', methods=['GET'])
@jwt_required()
@role_required('admin')
def get_active_sessions():
    """Get currently active sessions with filters"""
    db = get_db()
    
    # Get query parameters for filtering
    instructor_id = request.args.get('instructor_id')
    course_name = request.args.get('course_name')
    session_type = request.args.get('session_type')
    time_block = request.args.get('time_block')
    
    # Build query
    query = {'status': 'active'}
    
    if instructor_id:
        query['instructor_id'] = instructor_id
    if course_name:
        query['course_name'] = course_name
    if session_type and session_type in ['lab', 'theory']:
        query['session_type'] = session_type
    if time_block and time_block in ['morning', 'afternoon']:
        query['time_block'] = time_block
    
    active_sessions = list(db.sessions.find(query).sort('start_time', -1))
    
    sessions = []
    for session in active_sessions:
        instructor = db.users.find_one({'_id': ObjectId(session.get('instructor_id', ''))}) if session.get('instructor_id') else None
        
        sessions.append({
            'id': str(session['_id']),
            'name': session['name'],
            'instructor_name': instructor['name'] if instructor else 'Unknown',
            'instructor_id': session.get('instructor_id', ''),
            'course_name': session.get('course_name', ''),
            'session_type': session.get('session_type', ''),
            'section_id': session.get('section_id', ''),
            'year': session.get('year', ''),
            'time_block': session.get('time_block', ''),
            'start_time': session['start_time'].isoformat(),
            'attendance_count': session.get('attendance_count', 0),
            'status': 'active'
        })
    
    return jsonify(sessions), 200


@admin_bp.route('/recent-sessions', methods=['GET'])
@jwt_required()
@role_required('admin')
def get_recent_sessions():
    """Get recent completed sessions with filters"""
    db = get_db()
    
    # Get query parameters for filtering
    instructor_id = request.args.get('instructor_id')
    course_name = request.args.get('course_name')
    session_type = request.args.get('session_type')
    time_block = request.args.get('time_block')
    limit = int(request.args.get('limit', 50))
    
    # Build query
    query = {'status': 'completed'}
    
    if instructor_id:
        query['instructor_id'] = instructor_id
    if course_name:
        query['course_name'] = course_name
    if session_type and session_type in ['lab', 'theory']:
        query['session_type'] = session_type
    if time_block and time_block in ['morning', 'afternoon']:
        query['time_block'] = time_block
    
    recent_sessions = list(db.sessions.find(query).sort('start_time', -1).limit(limit))
    
    sessions = []
    for session in recent_sessions:
        instructor = db.users.find_one({'_id': ObjectId(session.get('instructor_id', ''))}) if session.get('instructor_id') else None
        
        sessions.append({
            'id': str(session['_id']),
            'name': session['name'],
            'instructor_name': instructor['name'] if instructor else 'Unknown',
            'instructor_id': session.get('instructor_id', ''),
            'course_name': session.get('course_name', ''),
            'session_type': session.get('session_type', ''),
            'section_id': session.get('section_id', ''),
            'year': session.get('year', ''),
            'time_block': session.get('time_block', ''),
            'start_time': session['start_time'].isoformat(),
            'end_time': session['end_time'].isoformat() if session.get('end_time') else None,
            'attendance_count': session.get('attendance_count', 0),
            'status': 'completed'
        })
    
    return jsonify(sessions), 200


@admin_bp.route('/instructor/<instructor_id>/toggle', methods=['PUT'])
@jwt_required()
@role_required('admin')
def toggle_instructor(instructor_id):
    """Enable/Disable an instructor (admin only)"""
    try:
        db = get_db()
        
        instructor = db.users.find_one({'_id': ObjectId(instructor_id), 'role': 'instructor'})
        
        if not instructor:
            return jsonify({'error': 'Instructor not found'}), 404
        
        # Toggle enabled status
        new_status = not instructor.get('enabled', True)
        
        db.users.update_one(
            {'_id': ObjectId(instructor_id)},
            {'$set': {'enabled': new_status}}
        )
        
        status_text = 'enabled' if new_status else 'disabled'
        print(f"✅ Instructor {instructor_id} {status_text}")
        
        return jsonify({
            'message': f'Instructor {status_text} successfully',
            'enabled': new_status
        }), 200
        
    except Exception as e:
        print(f"❌ Error toggling instructor: {e}")
        return jsonify({'error': 'Failed to toggle instructor'}), 500


@admin_bp.route('/student/<student_id>/toggle', methods=['PUT'])
@jwt_required()
@role_required('admin')
def toggle_student(student_id):
    """Enable/Disable a student (admin only)"""
    try:
        db = get_db()
        
        student = db.students.find_one({'_id': ObjectId(student_id)})
        
        if not student:
            return jsonify({'error': 'Student not found'}), 404
        
        # Get user_id and toggle in users collection
        user = db.users.find_one({'_id': ObjectId(student['user_id'])})
        
        if not user:
            return jsonify({'error': 'User not found'}), 404
        
        # Toggle enabled status
        new_status = not user.get('enabled', True)
        
        db.users.update_one(
            {'_id': ObjectId(student['user_id'])},
            {'$set': {'enabled': new_status}}
        )
        
        status_text = 'enabled' if new_status else 'disabled'
        print(f"✅ Student {student_id} {status_text}")
        
        return jsonify({
            'message': f'Student {status_text} successfully',
            'enabled': new_status
        }), 200
        
    except Exception as e:
        print(f"❌ Error toggling student: {e}")
        return jsonify({'error': 'Failed to toggle student'}), 500


@admin_bp.route('/instructor/<instructor_id>', methods=['PUT'])
@jwt_required()
@role_required('admin')
def update_instructor(instructor_id):
    """Update instructor details (admin only)"""
    try:
        data = request.get_json()
        db = get_db()
        
        instructor = db.users.find_one({'_id': ObjectId(instructor_id), 'role': 'instructor'})
        
        if not instructor:
            return jsonify({'error': 'Instructor not found'}), 404
        
        # Build update document
        update_doc = {}
        if 'name' in data:
            update_doc['name'] = data['name']
        if 'email' in data:
            update_doc['email'] = data['email']
        if 'department' in data:
            update_doc['department'] = data['department']
        
        if update_doc:
            db.users.update_one(
                {'_id': ObjectId(instructor_id)},
                {'$set': update_doc}
            )
        
        print(f"✅ Instructor {instructor_id} updated")
        return jsonify({'message': 'Instructor updated successfully'}), 200
        
    except Exception as e:
        print(f"❌ Error updating instructor: {e}")
        return jsonify({'error': 'Failed to update instructor'}), 500


@admin_bp.route('/student/<student_id>', methods=['PUT'])
@jwt_required()
@role_required('admin')
def update_student(student_id):
    """Update student details (admin only)"""
    try:
        data = request.get_json()
        db = get_db()
        
        student = db.students.find_one({'_id': ObjectId(student_id)})
        
        if not student:
            return jsonify({'error': 'Student not found'}), 404
        
        # Update student document
        student_update = {}
        if 'name' in data:
            student_update['name'] = data['name']
        if 'email' in data:
            student_update['email'] = data['email']
        if 'department' in data:
            student_update['department'] = data['department']
        if 'year' in data:
            student_update['year'] = data['year']
        if 'section' in data:
            student_update['section'] = data['section']
        
        if student_update:
            db.students.update_one(
                {'_id': ObjectId(student_id)},
                {'$set': student_update}
            )
        
        # Update user document
        user_update = {}
        if 'name' in data:
            user_update['name'] = data['name']
        if 'email' in data:
            user_update['email'] = data['email']
        
        if user_update:
            db.users.update_one(
                {'_id': ObjectId(student['user_id'])},
                {'$set': user_update}
            )
        
        print(f"✅ Student {student_id} updated")
        return jsonify({'message': 'Student updated successfully'}), 200
        
    except Exception as e:
        print(f"❌ Error updating student: {e}")
        return jsonify({'error': 'Failed to update student'}), 500
