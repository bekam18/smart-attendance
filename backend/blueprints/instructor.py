"""
Instructor-specific endpoints for records, settings, and exports
"""

from flask import Blueprint, request, jsonify, send_file
from flask_jwt_extended import jwt_required, get_jwt_identity
from datetime import datetime
from utils.security import role_required, hash_password, verify_password
from db.mysql import get_db
import csv
import io
import logging

instructor_bp = Blueprint('instructor', __name__)
logger = logging.getLogger(__name__)


@instructor_bp.route('/records', methods=['GET'])
@jwt_required()
@role_required('instructor', 'admin')
def get_attendance_records():
    """
    Get attendance records with filtering - instructors see ONLY their records
    Query params: start_date, end_date, student_id, session_id
    """
    try:
        user_id = get_jwt_identity()
        db = get_db()
        
        # Get query parameters
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')
        student_id = request.args.get('student_id')
        session_id = request.args.get('session_id')
        section_id = request.args.get('section_id')
        
        # Get user info
        user_result = db.execute_query('SELECT * FROM users WHERE id = %s', (user_id,))
        if not user_result:
            return jsonify({'error': 'User not found'}), 404
        user = user_result[0]
        
        # Build SQL query with filters
        sql = 'SELECT * FROM attendance WHERE 1=1'
        params = []
        
        # CRITICAL: Filter by instructor_id for instructors (admins see all)
        if user['role'] == 'instructor':
            sql += ' AND instructor_id = %s'
            params.append(user_id)
            logger.info(f"Instructor {user_id} - filtering by instructor_id")
        
        # Date range filter
        if start_date:
            sql += ' AND date >= %s'
            params.append(start_date)
        if end_date:
            sql += ' AND date <= %s'
            params.append(end_date)
        
        # Student filter
        if student_id:
            sql += ' AND student_id = %s'
            params.append(student_id)
        
        # Session filter
        if session_id:
            sql += ' AND session_id = %s'
            params.append(session_id)
        
        # Section filter
        if section_id:
            sql += ' AND section_id = %s'
            params.append(section_id)
        
        sql += ' ORDER BY timestamp DESC LIMIT 1000'
        
        logger.info(f"Fetching records with SQL: {sql}, params: {params}")
        
        # Get records
        records = db.execute_query(sql, tuple(params) if params else None)
        
        result = []
        for record in records:
            student_result = db.execute_query('SELECT * FROM students WHERE student_id = %s', (record['student_id'],))
            student = student_result[0] if student_result else None
            session_result = db.execute_query('SELECT * FROM sessions WHERE id = %s', (record['session_id'],))
            session = session_result[0] if session_result else None
            
            result.append({
                'id': str(record['id']),  # Use 'id' not '_id'
                'student_id': record['student_id'],
                'student_name': student['name'] if student else 'Unknown',
                'session_id': str(record['session_id']),
                'session_name': session['name'] if session else 'Unknown',
                'section_id': record.get('section_id', ''),
                'year': record.get('year', ''),
                'course_name': record.get('course_name', ''),
                'session_type': record.get('session_type', ''),
                'date': str(record['date']),
                'timestamp': record['timestamp'].isoformat() if record.get('timestamp') else None,
                'confidence': float(record.get('confidence', 0)) if record.get('confidence') else 0,
                'status': record.get('status', 'present')
            })
        
        logger.info(f"Returning {len(result)} records for user {user_id}")
        return jsonify(result), 200
    
    except Exception as e:
        logger.error(f"Error fetching records: {e}", exc_info=True)
        import traceback
        traceback.print_exc()
        return jsonify({'error': 'Failed to fetch records', 'message': str(e)}), 500


@instructor_bp.route('/records/export/csv', methods=['GET'])
@jwt_required()
@role_required('instructor', 'admin')
def export_csv():
    """Export attendance records to CSV - instructors see ONLY their records"""
    try:
        user_id = get_jwt_identity()
        db = get_db()
        
        # Get same filters
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')
        student_id = request.args.get('student_id')
        session_id = request.args.get('session_id')
        section_id = request.args.get('section_id')
        
        # Get user info
        user_result = db.execute_query('SELECT * FROM users WHERE id = %s', (user_id,))
        if not user_result:
            return jsonify({'error': 'User not found'}), 404
        user = user_result[0]
        
        # Build SQL query with filters (same as get_attendance_records)
        sql = 'SELECT * FROM attendance WHERE 1=1'
        params = []
        
        if user['role'] == 'instructor':
            sql += ' AND instructor_id = %s'
            params.append(user_id)
        
        if start_date:
            sql += ' AND date >= %s'
            params.append(start_date)
        if end_date:
            sql += ' AND date <= %s'
            params.append(end_date)
        
        if student_id:
            sql += ' AND student_id = %s'
            params.append(student_id)
        if session_id:
            sql += ' AND session_id = %s'
            params.append(session_id)
        
        if section_id:
            sql += ' AND section_id = %s'
            params.append(section_id)
        
        sql += ' ORDER BY timestamp DESC LIMIT 1000'
        
        # Get records
        records = db.execute_query(sql, tuple(params) if params else None)
        
        # Create CSV in memory
        output = io.StringIO()
        writer = csv.writer(output)
        
        # Write header
        writer.writerow(['Date', 'Time', 'Student ID', 'Student Name', 'Session', 'Confidence', 'Status'])
        
        # Write data
        for record in records:
            student_result = db.execute_query('SELECT * FROM students WHERE student_id = %s', (record['student_id'],))
            student = student_result[0] if student_result else None
            session_result = db.execute_query('SELECT * FROM sessions WHERE id = %s', (record['session_id'],))
            session = session_result[0] if session_result else None
            
            writer.writerow([
                record['date'],
                record['timestamp'].strftime('%H:%M:%S'),
                record['student_id'],
                student['name'] if student else 'Unknown',
                session['name'] if session else 'Unknown',
                f"{record.get('confidence', 0):.2%}",
                record.get('status', 'present')
            ])
        
        # Create response
        output.seek(0)
        csv_data = output.getvalue()
        
        # Create BytesIO for send_file
        csv_bytes = io.BytesIO(csv_data.encode('utf-8'))
        csv_bytes.seek(0)
        
        filename = f'attendance_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv'
        
        return send_file(
            csv_bytes,
            mimetype='text/csv',
            as_attachment=True,
            download_name=filename
        )
    
    except Exception as e:
        logger.error(f"Error exporting CSV: {e}", exc_info=True)
        return jsonify({'error': 'Failed to export CSV', 'message': str(e)}), 500


@instructor_bp.route('/records/export/excel', methods=['GET'])
@jwt_required()
@role_required('instructor', 'admin')
def export_excel():
    """Export attendance records to Excel - instructors see ONLY their records"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill
        
        user_id = get_jwt_identity()
        db = get_db()
        
        # Get same filters
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')
        student_id = request.args.get('student_id')
        session_id = request.args.get('session_id')
        section_id = request.args.get('section_id')
        
        # Get user info
        user_result = db.execute_query('SELECT * FROM users WHERE id = %s', (user_id,))
        if not user_result:
            return jsonify({'error': 'User not found'}), 404
        user = user_result[0]
        
        # Build SQL query with filters (same as get_attendance_records)
        sql = 'SELECT * FROM attendance WHERE 1=1'
        params = []
        
        if user['role'] == 'instructor':
            sql += ' AND instructor_id = %s'
            params.append(user_id)
        
        if start_date:
            sql += ' AND date >= %s'
            params.append(start_date)
        if end_date:
            sql += ' AND date <= %s'
            params.append(end_date)
        
        if student_id:
            sql += ' AND student_id = %s'
            params.append(student_id)
        if session_id:
            sql += ' AND session_id = %s'
            params.append(session_id)
        
        if section_id:
            sql += ' AND section_id = %s'
            params.append(section_id)
        
        sql += ' ORDER BY timestamp DESC LIMIT 1000'
        
        # Get records
        records = db.execute_query(sql, tuple(params) if params else None)
        
        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Attendance Records"
        
        # Style header
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        # Write header
        headers = ['Date', 'Time', 'Student ID', 'Student Name', 'Session', 'Confidence', 'Status']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
        
        # Write data
        for row_idx, record in enumerate(records, 2):
            student_result = db.execute_query('SELECT * FROM students WHERE student_id = %s', (record['student_id'],))
            student = student_result[0] if student_result else None
            session_result = db.execute_query('SELECT * FROM sessions WHERE id = %s', (record['session_id'],))
            session = session_result[0] if session_result else None
            
            ws.cell(row=row_idx, column=1, value=record['date'])
            ws.cell(row=row_idx, column=2, value=record['timestamp'].strftime('%H:%M:%S'))
            ws.cell(row=row_idx, column=3, value=record['student_id'])
            ws.cell(row=row_idx, column=4, value=student['name'] if student else 'Unknown')
            ws.cell(row=row_idx, column=5, value=session['name'] if session else 'Unknown')
            ws.cell(row=row_idx, column=6, value=f"{record.get('confidence', 0):.2%}")
            ws.cell(row=row_idx, column=7, value=record.get('status', 'present'))
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to BytesIO
        excel_bytes = io.BytesIO()
        wb.save(excel_bytes)
        excel_bytes.seek(0)
        
        filename = f'attendance_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        
        return send_file(
            excel_bytes,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
    
    except ImportError:
        return jsonify({'error': 'openpyxl not installed', 'message': 'Install with: pip install openpyxl'}), 500
    except Exception as e:
        logger.error(f"Error exporting Excel: {e}", exc_info=True)
        return jsonify({'error': 'Failed to export Excel', 'message': str(e)}), 500


@instructor_bp.route('/settings', methods=['GET'])
@jwt_required()
@role_required('instructor')
def get_settings():
    """Get instructor settings"""
    try:
        user_id = get_jwt_identity()
        db = get_db()
        
        settings_result = db.execute_query('SELECT * FROM user_settings WHERE user_id = %s', (user_id,))
        settings = settings_result[0] if settings_result else None
        
        if not settings:
            # Return defaults
            return jsonify({
                'confidence_threshold': 0.60,
                'capture_interval': 2,
                'auto_capture': True
            }), 200
        
        return jsonify({
            'confidence_threshold': settings.get('confidence_threshold', 0.60),
            'capture_interval': settings.get('capture_interval', 2),
            'auto_capture': settings.get('auto_capture', True)
        }), 200
    
    except Exception as e:
        logger.error(f"Error fetching settings: {e}", exc_info=True)
        return jsonify({'error': 'Failed to fetch settings', 'message': str(e)}), 500


@instructor_bp.route('/settings', methods=['PUT'])
@jwt_required()
@role_required('instructor')
def update_settings():
    """Update instructor settings"""
    try:
        user_id = get_jwt_identity()
        data = request.get_json()
        db = get_db()
        
        settings_doc = {
            'user_id': user_id,
            'confidence_threshold': float(data.get('confidence_threshold', 0.60)),
            'capture_interval': int(data.get('capture_interval', 2)),
            'auto_capture': bool(data.get('auto_capture', True)),
            'updated_at': datetime.utcnow()
        }
        
        # Check if settings exist, then update or insert
        existing = db.execute_query('SELECT id FROM user_settings WHERE user_id = %s', (user_id,))
        
        if existing:
            db.execute_query(
                'UPDATE user_settings SET face_recognition_threshold = %s, updated_at = %s WHERE user_id = %s',
                (settings_doc['face_recognition_threshold'], settings_doc['updated_at'], user_id),
                fetch=False
            )
        else:
            db.execute_query(
                'INSERT INTO user_settings (user_id, face_recognition_threshold, updated_at) VALUES (%s, %s, %s)',
                (user_id, settings_doc['face_recognition_threshold'], settings_doc['updated_at']),
                fetch=False
            )
        
        logger.info(f"Settings updated for user {user_id}")
        return jsonify({'message': 'Settings updated successfully'}), 200
    
    except Exception as e:
        logger.error(f"Error updating settings: {e}", exc_info=True)
        return jsonify({'error': 'Failed to update settings', 'message': str(e)}), 500


@instructor_bp.route('/change-password', methods=['PUT'])
@jwt_required()
@role_required('instructor')
def change_password():
    """Change instructor password"""
    try:
        user_id = get_jwt_identity()
        data = request.get_json()
        db = get_db()
        
        current_password = data.get('current_password')
        new_password = data.get('new_password')
        
        if not current_password or not new_password:
            return jsonify({'error': 'Both current and new password required'}), 400
        
        if len(new_password) < 6:
            return jsonify({'error': 'New password must be at least 6 characters'}), 400
        
        user_result = db.execute_query('SELECT * FROM users WHERE id = %s', (user_id,))
        user = user_result[0] if user_result else None
        
        if not user:
            return jsonify({'error': 'User not found'}), 404
        
        if not verify_password(current_password, user['password']):
            return jsonify({'error': 'Current password is incorrect'}), 401
        
        db.execute_query(
            'UPDATE users SET password = %s WHERE id = %s',
            (hash_password(new_password), user_id),
            fetch=False
        )
        
        logger.info(f"Password changed for user {user_id}")
        return jsonify({'message': 'Password changed successfully'}), 200
    
    except Exception as e:
        logger.error(f"Error changing password: {e}", exc_info=True)
        return jsonify({'error': 'Failed to change password', 'message': str(e)}), 500


@instructor_bp.route('/sections', methods=['GET'])
@jwt_required()
@role_required('instructor')
def get_instructor_sections():
    """Get instructor's assigned sections"""
    try:
        user_id = get_jwt_identity()
        db = get_db()
        
        user_result = db.execute_query('SELECT * FROM users WHERE id = %s', (user_id,))
        
        if not user_result:
            return jsonify({'error': 'Instructor not found'}), 404
        
        user = user_result[0]  # Get the first result
        
        # Parse sections JSON
        import json
        sections = []
        if user.get('sections'):
            try:
                sections = json.loads(user['sections'])
            except (json.JSONDecodeError, TypeError):
                sections = []
        
        return jsonify({
            'sections': sections,
            'instructor_name': user.get('name', 'Unknown')
        }), 200
    
    except Exception as e:
        logger.error(f"Error fetching sections: {e}", exc_info=True)
        return jsonify({'error': 'Failed to fetch sections', 'message': str(e)}), 500


@instructor_bp.route('/info', methods=['GET'])
@jwt_required()
@role_required('instructor')
def get_instructor_info():
    """Get instructor's information including session types"""
    try:
        user_id = get_jwt_identity()
        db = get_db()
        
        user_result = db.execute_query('SELECT * FROM users WHERE id = %s', (user_id,))
        
        if not user_result:
            return jsonify({'error': 'Instructor not found'}), 404
        
        user = user_result[0]  # Get the first (and only) result
        
        # Parse JSON fields
        import json
        session_types = []
        sections = []
        courses = []
        
        if user.get('session_types'):
            try:
                session_types = json.loads(user['session_types'])
            except (json.JSONDecodeError, TypeError):
                session_types = []
        
        if user.get('sections'):
            try:
                sections = json.loads(user['sections'])
            except (json.JSONDecodeError, TypeError):
                sections = []
        
        # Parse courses array (new multi-course field)
        if user.get('courses'):
            try:
                courses = json.loads(user['courses'])
            except (json.JSONDecodeError, TypeError):
                courses = []
        
        # Backward compatibility: if no courses array, use course_name
        if not courses and user.get('course_name'):
            courses = [user['course_name']]
        
        return jsonify({
            'name': user.get('name', 'Unknown'),
            'email': user.get('email', ''),
            'department': user.get('department', ''),
            'course_name': user.get('course_name', ''),  # Keep for backward compatibility
            'courses': courses,  # New multi-course field
            'class_year': user.get('class_year', ''),
            'session_types': session_types,
            'sections': sections
        }), 200
    
    except Exception as e:
        logger.error(f"Error fetching instructor info: {e}", exc_info=True)
        return jsonify({'error': 'Failed to fetch instructor info', 'message': str(e)}), 500


@instructor_bp.route('/sections-by-course', methods=['GET'])
@jwt_required()
@role_required('instructor', 'admin')
def get_sections_by_course():
    """Get sections for a specific course that the instructor teaches"""
    try:
        user_id = get_jwt_identity()
        db = get_db()
        
        course_name = request.args.get('course_name')
        if not course_name:
            return jsonify({'error': 'course_name parameter is required'}), 400
        
        logger.info(f"Getting sections for instructor {user_id}, course: {course_name}")
        
        # Get sections from sessions table where this instructor has taught this course
        query = """
            SELECT DISTINCT section_id
            FROM sessions
            WHERE instructor_id = %s AND course_name = %s AND section_id IS NOT NULL
            ORDER BY section_id
        """
        
        result = db.execute_query(query, (user_id, course_name))
        sections = [row['section_id'] for row in result]
        
        # If no sections found in sessions, try attendance table
        if not sections:
            query = """
                SELECT DISTINCT section_id
                FROM attendance
                WHERE instructor_id = %s AND course_name = %s AND section_id IS NOT NULL
                ORDER BY section_id
            """
            result = db.execute_query(query, (user_id, course_name))
            sections = [row['section_id'] for row in result]
        
        logger.info(f"Found sections for {course_name}: {sections}")
        
        return jsonify({'sections': sections}), 200
    
    except Exception as e:
        logger.error(f"Error fetching sections by course: {e}", exc_info=True)
        return jsonify({'error': 'Failed to fetch sections', 'message': str(e)}), 500


@instructor_bp.route('/students', methods=['GET'])
@jwt_required()
@role_required('instructor', 'admin')
def get_students_list():
    """Get list of students - instructors see only students from their attendance records"""
    try:
        user_id = get_jwt_identity()
        db = get_db()
        
        user_result = db.execute_query('SELECT * FROM users WHERE id = %s', (user_id,))
        user = user_result[0] if user_result else None
        
        if not user:
            return jsonify({'error': 'User not found'}), 404
        
        if user['role'] == 'instructor':
            # Get unique student IDs from instructor's attendance records
            attendance_records = db.execute_query('SELECT DISTINCT student_id FROM attendance WHERE instructor_id = %s', (user_id,))
            student_ids = [record['student_id'] for record in attendance_records]
            
            if student_ids:
                # Create placeholders for IN clause
                placeholders = ','.join(['%s'] * len(student_ids))
                students = db.execute_query(f'SELECT student_id, name FROM students WHERE student_id IN ({placeholders}) ORDER BY student_id', student_ids)
            else:
                students = []
        else:
            # Admins see all students
            students = db.execute_query('SELECT student_id, name FROM students ORDER BY student_id')
        
        result = [
            {
                'student_id': s['student_id'],
                'name': s['name']
            }
            for s in students
        ]
        
        return jsonify(result), 200
    
    except Exception as e:
        logger.error(f"Error fetching students: {e}", exc_info=True)
        return jsonify({'error': 'Failed to fetch students', 'message': str(e)}), 500


@instructor_bp.route('/reports/generate', methods=['POST'])
@jwt_required()
@role_required('instructor', 'admin')
def generate_report():
    """Generate comprehensive attendance report with statistics"""
    try:
        user_id = get_jwt_identity()
        db = get_db()
        data = request.get_json()
        
        # Get user info
        user_result = db.execute_query('SELECT * FROM users WHERE id = %s', (user_id,))
        if not user_result:
            return jsonify({'error': 'User not found'}), 404
        user = user_result[0]
        
        # Extract filters
        report_type = data.get('report_type')  # daily, weekly, monthly, semester, yearly
        section_id = data.get('section_id')
        course_name = data.get('course_name')
        start_date = data.get('start_date')
        end_date = data.get('end_date')
        
        # Build query for attendance records
        sql = 'SELECT * FROM attendance WHERE 1=1'
        params = []
        
        # Filter by instructor for non-admins
        if user['role'] == 'instructor':
            sql += ' AND instructor_id = %s'
            params.append(user_id)
        
        # Apply filters
        if section_id:
            sql += ' AND section_id = %s'
            params.append(section_id)
        
        if course_name:
            sql += ' AND course_name = %s'
            params.append(course_name)
        
        if start_date:
            sql += ' AND date >= %s'
            params.append(start_date)
        
        if end_date:
            sql += ' AND date <= %s'
            params.append(end_date)
        
        sql += ' ORDER BY date, timestamp'
        
        # Get attendance records
        records = db.execute_query(sql, tuple(params) if params else None)
        
        # Get all students in the section
        student_sql = 'SELECT * FROM students WHERE 1=1'
        student_params = []
        
        if section_id:
            student_sql += ' AND section = %s'
            student_params.append(section_id)
        
        students = db.execute_query(student_sql, tuple(student_params) if student_params else None)
        
        # Calculate statistics per student
        student_stats = {}
        session_ids = set()
        session_types = {}  # Track session types by session_id
        
        # First pass: collect all unique sessions and their types
        for record in records:
            session_id = record.get('session_id')
            if session_id:
                session_ids.add(session_id)
                session_types[session_id] = record.get('session_type', 'theory')
        
        # Second pass: count attendance per student
        logger.info(f"Processing {len(records)} attendance records")
        for record in records:
            student_id = record['student_id']
            session_id = record.get('session_id')
            session_type = record.get('session_type', 'theory')
            status = record.get('status')
            
            logger.info(f"Record: student={student_id}, session={session_id}, status={status}, type={session_type}")
            
            if student_id not in student_stats:
                student_stats[student_id] = {
                    'student_id': student_id,
                    'name': '',
                    'section': record.get('section_id', ''),
                    'sessions_attended': set(),  # Track which sessions this student attended
                    'lab_sessions_attended': set(),
                    'theory_sessions_attended': set(),
                    'total_sessions': 0,
                    'present_count': 0,
                    'absent_count': 0,
                    'lab_sessions': 0,
                    'lab_present': 0,
                    'theory_sessions': 0,
                    'theory_present': 0,
                    'percentage': 0,
                    'lab_percentage': 0,
                    'theory_percentage': 0,
                    'below_threshold': False
                }
            
            # Track session attendance
            if session_id:
                student_stats[student_id]['sessions_attended'].add(session_id)
                
                if status == 'present':
                    student_stats[student_id]['present_count'] += 1
                    logger.info(f"✅ Incremented present_count for {student_id}: now {student_stats[student_id]['present_count']}")
                    if session_type == 'lab':
                        student_stats[student_id]['lab_sessions_attended'].add(session_id)
                        student_stats[student_id]['lab_present'] += 1
                    else:
                        student_stats[student_id]['theory_sessions_attended'].add(session_id)
                        student_stats[student_id]['theory_present'] += 1
                # Note: absent_count will be calculated later as (total_sessions - present_count)
        
        # Count total lab and theory sessions
        total_lab_sessions = sum(1 for sid, stype in session_types.items() if stype == 'lab')
        total_theory_sessions = sum(1 for sid, stype in session_types.items() if stype == 'theory')
        
        # Add student names and calculate percentages
        for student in students:
            student_id = student['student_id']
            if student_id in student_stats:
                student_stats[student_id]['name'] = student['name']
                
                # Set total sessions based on session types
                student_stats[student_id]['lab_sessions'] = total_lab_sessions
                student_stats[student_id]['theory_sessions'] = total_theory_sessions
                student_stats[student_id]['total_sessions'] = len(session_ids)
                
                # Calculate absent count
                student_stats[student_id]['absent_count'] = (
                    student_stats[student_id]['total_sessions'] - 
                    student_stats[student_id]['present_count']
                )
            else:
                # Student with no attendance records (all absent)
                student_stats[student_id] = {
                    'student_id': student_id,
                    'name': student['name'],
                    'section': student.get('section', ''),
                    'total_sessions': len(session_ids),
                    'present_count': 0,
                    'absent_count': len(session_ids),
                    'lab_sessions': total_lab_sessions,
                    'lab_present': 0,
                    'theory_sessions': total_theory_sessions,
                    'theory_present': 0,
                    'percentage': 0,
                    'lab_percentage': 0,
                    'theory_percentage': 0,
                    'below_threshold': True
                }
        
        # Calculate percentages and thresholds
        for student_id, stats in student_stats.items():
            # Remove temporary sets
            stats.pop('sessions_attended', None)
            stats.pop('lab_sessions_attended', None)
            stats.pop('theory_sessions_attended', None)
            
            # Calculate overall percentage
            if stats['total_sessions'] > 0:
                stats['percentage'] = (stats['present_count'] / stats['total_sessions']) * 100
            
            # Calculate lab percentage
            if stats['lab_sessions'] > 0:
                stats['lab_percentage'] = (stats['lab_present'] / stats['lab_sessions']) * 100
                if stats['lab_percentage'] < 100:
                    stats['below_threshold'] = True
            
            # Calculate theory percentage
            if stats['theory_sessions'] > 0:
                stats['theory_percentage'] = (stats['theory_present'] / stats['theory_sessions']) * 100
                if stats['theory_percentage'] < 80:
                    stats['below_threshold'] = True
        
        # Convert to list and sort
        report_data = sorted(student_stats.values(), key=lambda x: x['student_id'])
        
        # Debug: Log final report data
        logger.info(f"Final report: {len(report_data)} students, {len(session_ids)} sessions")
        for student in report_data[:3]:  # Log first 3 students
            logger.info(f"Student {student['student_id']}: present={student['present_count']}, absent={student['absent_count']}")
        
        return jsonify({
            'report_type': report_type,
            'section_id': section_id,
            'course_name': course_name,
            'start_date': start_date,
            'end_date': end_date,
            'total_sessions': len(session_ids),
            'total_students': len(report_data),
            'data': report_data
        }), 200
    
    except Exception as e:
        logger.error(f"Error generating report: {e}", exc_info=True)
        return jsonify({'error': 'Failed to generate report', 'message': str(e)}), 500


@instructor_bp.route('/reports/download/csv', methods=['POST'])
@jwt_required()
@role_required('instructor', 'admin')
def download_report_csv():
    """Download attendance report as CSV"""
    try:
        user_id = get_jwt_identity()
        db = get_db()
        data = request.get_json()
        
        # Get user info
        user_result = db.execute_query('SELECT * FROM users WHERE id = %s', (user_id,))
        if not user_result:
            return jsonify({'error': 'User not found'}), 404
        user = user_result[0]
        
        # Extract filters (same as generate_report)
        report_type = data.get('report_type', 'custom')
        section_id = data.get('section_id')
        course_name = data.get('course_name')
        start_date = data.get('start_date')
        end_date = data.get('end_date')
        
        # Build query
        sql = 'SELECT * FROM attendance WHERE 1=1'
        params = []
        
        if user['role'] == 'instructor':
            sql += ' AND instructor_id = %s'
            params.append(user_id)
        
        if section_id:
            sql += ' AND section_id = %s'
            params.append(section_id)
        
        if course_name:
            sql += ' AND course_name = %s'
            params.append(course_name)
        
        if start_date:
            sql += ' AND date >= %s'
            params.append(start_date)
        
        if end_date:
            sql += ' AND date <= %s'
            params.append(end_date)
        
        sql += ' ORDER BY date, timestamp'
        
        records = db.execute_query(sql, tuple(params) if params else None)
        
        # Get students
        student_sql = 'SELECT * FROM students WHERE 1=1'
        student_params = []
        if section_id:
            student_sql += ' AND section = %s'
            student_params.append(section_id)
        students = db.execute_query(student_sql, tuple(student_params) if student_params else None)
        
        # Calculate statistics using CORRECT logic (same as generate_report)
        student_stats = {}
        session_ids = set()
        session_types = {}
        
        # First pass: collect unique sessions
        for record in records:
            session_id = record.get('session_id')
            if session_id:
                session_ids.add(session_id)
                session_types[session_id] = record.get('session_type', 'theory')
        
        # Second pass: count attendance per student
        for record in records:
            student_id = record['student_id']
            session_id = record.get('session_id')
            session_type = record.get('session_type', 'theory')
            status = record.get('status')
            
            if student_id not in student_stats:
                student_stats[student_id] = {
                    'student_id': student_id,
                    'name': '',
                    'section': record.get('section_id', ''),
                    'sessions_attended': set(),
                    'total_sessions': 0,
                    'present_count': 0,
                    'absent_count': 0,
                    'lab_sessions': 0,
                    'lab_present': 0,
                    'theory_sessions': 0,
                    'theory_present': 0
                }
            
            # Track session attendance
            if session_id:
                student_stats[student_id]['sessions_attended'].add(session_id)
                
                if status == 'present':
                    student_stats[student_id]['present_count'] += 1
                    if session_type == 'lab':
                        student_stats[student_id]['lab_present'] += 1
                    else:
                        student_stats[student_id]['theory_present'] += 1
        
        # Count total lab and theory sessions
        total_lab_sessions = sum(1 for sid, stype in session_types.items() if stype == 'lab')
        total_theory_sessions = sum(1 for sid, stype in session_types.items() if stype == 'theory')
        
        # Add student names and calculate
        for student in students:
            student_id = student['student_id']
            if student_id in student_stats:
                student_stats[student_id]['name'] = student['name']
                student_stats[student_id]['lab_sessions'] = total_lab_sessions
                student_stats[student_id]['theory_sessions'] = total_theory_sessions
                student_stats[student_id]['total_sessions'] = len(session_ids)
                student_stats[student_id]['absent_count'] = (
                    student_stats[student_id]['total_sessions'] - 
                    student_stats[student_id]['present_count']
                )
            else:
                # Student with no attendance records
                student_stats[student_id] = {
                    'student_id': student_id,
                    'name': student['name'],
                    'section': student.get('section', ''),
                    'total_sessions': len(session_ids),
                    'present_count': 0,
                    'absent_count': len(session_ids),
                    'lab_sessions': total_lab_sessions,
                    'lab_present': 0,
                    'theory_sessions': total_theory_sessions,
                    'theory_present': 0
                }
        
        # Remove temporary sets
        for stats in student_stats.values():
            stats.pop('sessions_attended', None)
        
        # Create CSV
        output = io.StringIO()
        writer = csv.writer(output)
        
        # Write header with report info
        writer.writerow([f'Attendance Report - {report_type.title()}'])
        writer.writerow([f'Section: {section_id or "All"}'])
        writer.writerow([f'Course: {course_name or "All"}'])
        writer.writerow([f'Period: {start_date or "Start"} to {end_date or "End"}'])
        writer.writerow([f'Total Sessions: {len(session_ids)}'])
        writer.writerow([])
        
        # Write data header
        writer.writerow([
            'Student ID', 'Name', 'Section', 
            'Total Sessions', 'Present', 'Absent', 'Overall %',
            'Lab Sessions', 'Lab Present', 'Lab %',
            'Theory Sessions', 'Theory Present', 'Theory %',
            'Below Threshold'
        ])
        
        # Write data
        for student_id in sorted(student_stats.keys()):
            stats = student_stats[student_id]
            
            overall_pct = (stats['present_count'] / stats['total_sessions'] * 100) if stats['total_sessions'] > 0 else 0
            lab_pct = (stats['lab_present'] / stats['lab_sessions'] * 100) if stats['lab_sessions'] > 0 else 0
            theory_pct = (stats['theory_present'] / stats['theory_sessions'] * 100) if stats['theory_sessions'] > 0 else 0
            
            below_threshold = (lab_pct < 100 and stats['lab_sessions'] > 0) or (theory_pct < 80 and stats['theory_sessions'] > 0)
            
            writer.writerow([
                stats['student_id'],
                stats['name'],
                stats['section'],
                stats['total_sessions'],
                stats['present_count'],
                stats['absent_count'],
                f"{overall_pct:.1f}%",
                stats['lab_sessions'],
                stats['lab_present'],
                f"{lab_pct:.1f}%",
                stats['theory_sessions'],
                stats['theory_present'],
                f"{theory_pct:.1f}%",
                'YES' if below_threshold else 'NO'
            ])
        
        output.seek(0)
        csv_data = output.getvalue()
        csv_bytes = io.BytesIO(csv_data.encode('utf-8'))
        csv_bytes.seek(0)
        
        filename = f'attendance_report_{report_type}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv'
        
        return send_file(
            csv_bytes,
            mimetype='text/csv',
            as_attachment=True,
            download_name=filename
        )
    
    except Exception as e:
        logger.error(f"Error downloading CSV report: {e}", exc_info=True)
        return jsonify({'error': 'Failed to download CSV', 'message': str(e)}), 500


@instructor_bp.route('/reports/download/excel', methods=['POST'])
@jwt_required()
@role_required('instructor', 'admin')
def download_report_excel():
    """Download attendance report as Excel with formatting"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        
        user_id = get_jwt_identity()
        db = get_db()
        data = request.get_json()
        
        # Get user info
        user_result = db.execute_query('SELECT * FROM users WHERE id = %s', (user_id,))
        if not user_result:
            return jsonify({'error': 'User not found'}), 404
        user = user_result[0]
        
        # Extract filters
        report_type = data.get('report_type', 'custom')
        section_id = data.get('section_id')
        course_name = data.get('course_name')
        start_date = data.get('start_date')
        end_date = data.get('end_date')
        
        # Build query
        sql = 'SELECT * FROM attendance WHERE 1=1'
        params = []
        
        if user['role'] == 'instructor':
            sql += ' AND instructor_id = %s'
            params.append(user_id)
        
        if section_id:
            sql += ' AND section_id = %s'
            params.append(section_id)
        
        if course_name:
            sql += ' AND course_name = %s'
            params.append(course_name)
        
        if start_date:
            sql += ' AND date >= %s'
            params.append(start_date)
        
        if end_date:
            sql += ' AND date <= %s'
            params.append(end_date)
        
        sql += ' ORDER BY date, timestamp'
        
        records = db.execute_query(sql, tuple(params) if params else None)
        
        # Get students
        student_sql = 'SELECT * FROM students WHERE 1=1'
        student_params = []
        if section_id:
            student_sql += ' AND section = %s'
            student_params.append(section_id)
        students = db.execute_query(student_sql, tuple(student_params) if student_params else None)
        
        # Calculate statistics using CORRECT logic (same as generate_report)
        student_stats = {}
        session_ids = set()
        session_types = {}
        
        # First pass: collect unique sessions
        for record in records:
            session_id = record.get('session_id')
            if session_id:
                session_ids.add(session_id)
                session_types[session_id] = record.get('session_type', 'theory')
        
        # Second pass: count attendance per student
        for record in records:
            student_id = record['student_id']
            session_id = record.get('session_id')
            session_type = record.get('session_type', 'theory')
            status = record.get('status')
            
            if student_id not in student_stats:
                student_stats[student_id] = {
                    'student_id': student_id,
                    'name': '',
                    'section': record.get('section_id', ''),
                    'sessions_attended': set(),
                    'total_sessions': 0,
                    'present_count': 0,
                    'absent_count': 0,
                    'lab_sessions': 0,
                    'lab_present': 0,
                    'theory_sessions': 0,
                    'theory_present': 0
                }
            
            # Track session attendance
            if session_id:
                student_stats[student_id]['sessions_attended'].add(session_id)
                
                if status == 'present':
                    student_stats[student_id]['present_count'] += 1
                    if session_type == 'lab':
                        student_stats[student_id]['lab_present'] += 1
                    else:
                        student_stats[student_id]['theory_present'] += 1
        
        # Count total lab and theory sessions
        total_lab_sessions = sum(1 for sid, stype in session_types.items() if stype == 'lab')
        total_theory_sessions = sum(1 for sid, stype in session_types.items() if stype == 'theory')
        
        # Add student names and calculate
        for student in students:
            student_id = student['student_id']
            if student_id in student_stats:
                student_stats[student_id]['name'] = student['name']
                student_stats[student_id]['lab_sessions'] = total_lab_sessions
                student_stats[student_id]['theory_sessions'] = total_theory_sessions
                student_stats[student_id]['total_sessions'] = len(session_ids)
                student_stats[student_id]['absent_count'] = (
                    student_stats[student_id]['total_sessions'] - 
                    student_stats[student_id]['present_count']
                )
            else:
                # Student with no attendance records
                student_stats[student_id] = {
                    'student_id': student_id,
                    'name': student['name'],
                    'section': student.get('section', ''),
                    'total_sessions': len(session_ids),
                    'present_count': 0,
                    'absent_count': len(session_ids),
                    'lab_sessions': total_lab_sessions,
                    'lab_present': 0,
                    'theory_sessions': total_theory_sessions,
                    'theory_present': 0
                }
        
        # Remove temporary sets
        for stats in student_stats.values():
            stats.pop('sessions_attended', None)
        
        # Create Excel workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Attendance Report"
        
        # Styles
        title_font = Font(size=14, bold=True, color="FFFFFF")
        title_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        warning_fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Write title
        ws.merge_cells('A1:N1')
        title_cell = ws['A1']
        title_cell.value = f'Attendance Report - {report_type.title()}'
        title_cell.font = title_font
        title_cell.fill = title_fill
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Write report info
        ws['A2'] = f'Section: {section_id or "All"}'
        ws['A3'] = f'Course: {course_name or "All"}'
        ws['A4'] = f'Period: {start_date or "Start"} to {end_date or "End"}'
        ws['A5'] = f'Total Sessions: {len(session_ids)}'
        
        # Write header row
        headers = [
            'Student ID', 'Name', 'Section',
            'Total Sessions', 'Present', 'Absent', 'Overall %',
            'Lab Sessions', 'Lab Present', 'Lab %',
            'Theory Sessions', 'Theory Present', 'Theory %',
            'Below Threshold'
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=7, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
        
        # Write data
        row_idx = 8
        for student_id in sorted(student_stats.keys()):
            stats = student_stats[student_id]
            
            overall_pct = (stats['present_count'] / stats['total_sessions'] * 100) if stats['total_sessions'] > 0 else 0
            lab_pct = (stats['lab_present'] / stats['lab_sessions'] * 100) if stats['lab_sessions'] > 0 else 0
            theory_pct = (stats['theory_present'] / stats['theory_sessions'] * 100) if stats['theory_sessions'] > 0 else 0
            
            below_threshold = (lab_pct < 100 and stats['lab_sessions'] > 0) or (theory_pct < 80 and stats['theory_sessions'] > 0)
            
            row_data = [
                stats['student_id'],
                stats['name'],
                stats['section'],
                stats['total_sessions'],
                stats['present_count'],
                stats['absent_count'],
                f"{overall_pct:.1f}%",
                stats['lab_sessions'],
                stats['lab_present'],
                f"{lab_pct:.1f}%",
                stats['theory_sessions'],
                stats['theory_present'],
                f"{theory_pct:.1f}%",
                'YES' if below_threshold else 'NO'
            ]
            
            for col, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col, value=value)
                cell.border = border
                cell.alignment = Alignment(horizontal='center', vertical='center')
                
                # Highlight rows below threshold
                if below_threshold:
                    cell.fill = warning_fill
            
            row_idx += 1
        
        # Auto-adjust column widths
        for col_idx in range(1, len(headers) + 1):
            max_length = 0
            column_letter = get_column_letter(col_idx)
            
            for row in ws.iter_rows(min_col=col_idx, max_col=col_idx):
                for cell in row:
                    try:
                        if cell.value and len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
            
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to BytesIO
        excel_bytes = io.BytesIO()
        wb.save(excel_bytes)
        excel_bytes.seek(0)
        
        filename = f'attendance_report_{report_type}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        
        return send_file(
            excel_bytes,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
    
    except ImportError:
        return jsonify({'error': 'openpyxl not installed', 'message': 'Install with: pip install openpyxl'}), 500
    except Exception as e:
        logger.error(f"Error downloading Excel report: {e}", exc_info=True)
        return jsonify({'error': 'Failed to download Excel', 'message': str(e)}), 500
